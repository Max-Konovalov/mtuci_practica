from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os

# Регистрация кириллического шрифта для PDF
FONT_PATH = os.path.join('static', 'fonts', 'DejaVuSans.ttf')
pdfmetrics.registerFont(TTFont('DejaVu', FONT_PATH))

# Константы
LOGO_PATH = os.path.join('static', 'logo.png')
DEFAULT_FONT = 'DejaVu'
PDF_HEADER_SIZE = 20
PDF_NORMAL_SIZE = 12
PDF_SMALL_SIZE = 10
PDF_MARGIN = 50


def generate_pdf(count: int) -> str:
    """
    Генерирует PDF-отчёт с количеством найденных чемоданов.
    
    :param count: количество чемоданов
    :return: путь к файлу отчёта
    """
    file_path = 'static/report.pdf'
    c = canvas.Canvas(file_path, pagesize=A4)
    width, height = A4

    # Заголовок
    c.setFont(DEFAULT_FONT, PDF_HEADER_SIZE)
    c.drawCentredString(width / 2, height - 80, "Отчёт по подсчёту багажа")

    # Дата и время
    now = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
    c.setFont(DEFAULT_FONT, PDF_NORMAL_SIZE)
    c.drawRightString(width - PDF_MARGIN, height - 50, f"Дата: {now}")

    # Логотип
    if os.path.exists(LOGO_PATH):
        c.drawImage(LOGO_PATH, PDF_MARGIN, height - 120, width=80,
                    preserveAspectRatio=True, mask='auto')

    # Горизонтальная линия
    c.setStrokeColorRGB(0.5, 0.5, 0.5)  # Серый цвет
    c.line(PDF_MARGIN, height - 100, width - PDF_MARGIN, height - 100)

    # Основной текст
    c.setFont(DEFAULT_FONT, 14)
    c.drawString(100, height - 180, f"📦 Обнаружено чемоданов: {count}")

    # Подвал
    c.setFont(DEFAULT_FONT, PDF_SMALL_SIZE)
    c.setFillColorRGB(0.5, 0.5, 0.5)  # Серый цвет
    c.drawCentredString(width / 2, 40, "Система мониторинга багажа © 2025")

    c.save()
    return file_path


def generate_excel(count: int) -> str:
    """
    Генерирует Excel-файл с количеством найденных чемоданов.
    
    :param count: количество чемоданов
    :return: путь к файлу отчёта
    """
    now = datetime.now()
    now_str = now.strftime('%Y-%m-%d_%H-%M-%S')
    reports_dir = 'reports'

    os.makedirs(reports_dir, exist_ok=True)
    filename = f'{reports_dir}/report_{now_str}.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    # Заголовки таблицы
    ws.append(['Дата и время', 'Количество багажа'])

    # Стиль заголовков
    header_font = Font(bold=True)
    for col in range(1, 3):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = 25

    # Данные
    ws.append([now.strftime('%d.%m.%Y %H:%M:%S'), count])

    wb.save(filename)
    return filename